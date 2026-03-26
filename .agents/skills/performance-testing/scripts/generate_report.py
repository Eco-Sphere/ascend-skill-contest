#!/usr/bin/env python3
"""生成性能测试Excel报告"""

import json
import glob
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def load_results(result_dir):
    """加载测试结果"""
    results = []
    
    if not os.path.exists(result_dir):
        print(f"错误: 目录不存在 - {result_dir}")
        return results
    
    pattern = os.path.join(result_dir, "**", "results.json")
    files = glob.glob(pattern, recursive=True)
    
    if not files:
        print(f"警告: 未找到results.json文件 - {result_dir}")
        return results
    
    for f in files:
        try:
            with open(f, 'r', encoding='utf-8') as fp:
                data = json.load(fp)
                metrics = data.get('result_metrics', {})
                
                # 获取迭代目录名
                iter_dir = os.path.dirname(f)
                iter_name = os.path.basename(iter_dir)
                
                r = {
                    'iteration': iter_name,
                    'concurrency': data.get('concurrency', 1),
                    'request_rate': data.get('request_rate', 0),
                    
                    # TTFT 指标
                    'ttft_mean': float(metrics.get('mean_ttft_ms', 0)),
                    'ttft_median': float(metrics.get('median_ttft_ms', 0)),
                    'ttft_p99': float(metrics.get('p99_ttft_ms', 0)),
                    'ttft_max': float(metrics.get('max_ttft_ms', 0)),
                    
                    # TPOT 指标
                    'tpot_mean': float(metrics.get('mean_tpot_ms', 0)),
                    'tpot_median': float(metrics.get('median_tpot_ms', 0)),
                    'tpot_p99': float(metrics.get('p99_tpot_ms', 0)),
                    'tpot_max': float(metrics.get('max_tpot_ms', 0)),
                    
                    # 吞吐量指标
                    'qps': float(metrics.get('qps', 0)),
                    'request_throughput': float(metrics.get('request_throughput', 0)),
                    'input_throughput': float(metrics.get('input_throughput', 0)),
                    'output_throughput': float(metrics.get('output_throughput', 0)),
                    'total_token_throughput': float(metrics.get('total_token_throughput', 0)),
                    
                    # 请求延迟
                    'latency_mean': float(metrics.get('latency_mean_ms', 0)),
                    'latency_p50': float(metrics.get('latency_p50_ms', 0)),
                    'latency_p99': float(metrics.get('latency_p99_ms', 0)),
                    
                    # 错误统计
                    'failures': int(metrics.get('failures', 0)),
                    'total_requests': int(data.get('num_prompts', 0)),
                }
                results.append(r)
                
        except json.JSONDecodeError as e:
            print(f"JSON解析失败 {f}: {e}")
        except Exception as e:
            print(f"加载失败 {f}: {e}")
    
    # 按并发和请求频率排序
    return sorted(results, key=lambda x: (x['concurrency'], x['request_rate']))


def generate_excel_report(results, best_config, target_ttft, target_tpot, output_file="benchmark_report.xlsx"):
    """生成Excel报告"""
    if not results:
        print("错误: 没有数据可导出")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "性能测试报告"
    
    title_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='微软雅黑', size=10, bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    cell_font = Font(name='微软雅黑', size=9)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 标题
    ws.merge_cells('A1:R1')
    ws['A1'] = '推理服务性能测试报告'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # 测试信息
    info_rows = [
        ('A3', '测试时间', 'B3', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('A4', 'TTFT目标', 'B4', f'<{target_ttft}ms'),
        ('A5', 'TPOT目标', 'B5', f'<{target_tpot}ms'),
    ]
    for cell_a, text_a, cell_b, text_b in info_rows:
        ws[cell_a] = text_a
        ws[cell_a].font = header_font
        ws[cell_b] = text_b
        ws[cell_b].font = cell_font
    
    # 表头
    headers = [
        '迭代号', '并发', '请求频率',
        'TTFT均值', 'TTFT中位数', 'TTFT P99', 'TTFT最大',
        'TPOT均值', 'TPOT中位数', 'TPOT P99', 'TPOT最大',
        'QPS', '请求吞吐', '输入吞吐', '输出吞吐', '总吞吐',
        '失败/总计', 'TTFT达标', 'TPOT达标', '状态'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # 数据
    for row_idx, r in enumerate(results, start=8):
        ttft_valid = r['ttft_mean'] < target_ttft
        tpot_valid = r['tpot_mean'] < target_tpot
        is_valid = ttft_valid and tpot_valid
        
        data = [
            r.get('iteration', ''),
            r['concurrency'],
            r['request_rate'],
            round(r['ttft_mean'], 1),
            round(r['ttft_median'], 1),
            round(r['ttft_p99'], 1),
            round(r['ttft_max'], 1),
            round(r['tpot_mean'], 1),
            round(r['tpot_median'], 1),
            round(r['tpot_p99'], 1),
            round(r['tpot_max'], 1),
            round(r['qps'], 2),
            round(r['request_throughput'], 2),
            round(r['input_throughput'], 2),
            round(r['output_throughput'], 2),
            round(r['total_token_throughput'], 2),
            f"{r['failures']}/{r['total_requests']}",
            '是' if ttft_valid else '否',
            '是' if tpot_valid else '否',
            '通过' if is_valid else '不通过',
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = cell_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # 状态列颜色
        fill_color = 'C6EFCE' if is_valid else 'FFC7CE'
        ws.cell(row=row_idx, column=len(headers)).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    
    # 最优配置
    if best_config:
        row = len(results) + 10
        ws.merge_cells(f'A{row}:R{row}')
        ws[f'A{row}'] = (
            f"最优配置：并发={best_config['concurrency']}, "
            f"请求频率={best_config['request_rate']}, "
            f"QPS={best_config['qps']:.2f}, "
            f"TTFT={best_config['ttft_mean']:.1f}ms, "
            f"TPOT={best_config['tpot_mean']:.1f}ms"
        )
        ws[f'A{row}'].font = Font(name='微软雅黑', size=11, bold=True, color='0070C0')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[row].height = 20
    
    # 列宽
    widths = [12, 8, 10, 10, 10, 10, 10, 10, 10, 10, 10, 8, 10, 10, 10, 10, 12, 10, 10, 10]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    try:
        wb.save(output_file)
        print(f"报告已生成: {output_file}")
    except Exception as e:
        print(f"保存文件失败: {e}")
        sys.exit(1)


def main():
    import argparse
    parser = argparse.ArgumentParser(description='生成性能测试Excel报告')
    parser.add_argument('--result-dir', default='./benchmark_results', help='结果目录')
    parser.add_argument('--output', default='benchmark_report.xlsx', help='输出文件')
    parser.add_argument('--target-ttft', type=float, default=2000, help='TTFT目标(ms)')
    parser.add_argument('--target-tpot', type=float, default=50, help='TPOT目标(ms)')
    args = parser.parse_args()
    
    print(f"加载测试结果: {args.result_dir}")
    results = load_results(args.result_dir)
    
    if not results:
        print("未找到测试结果")
        sys.exit(1)
    
    target_ttft = args.target_ttft
    target_tpot = args.target_tpot
    
    valid_results = [r for r in results if r['ttft_mean'] < target_ttft and r['tpot_mean'] < target_tpot]
    best_config = max(valid_results, key=lambda x: x['qps']) if valid_results else None
    
    print(f"测试结果: {len(valid_results)}/{len(results)} 通过")
    if best_config:
        print(f"最优配置: 并发={best_config['concurrency']}, 请求频率={best_config['request_rate']}, QPS={best_config['qps']:.2f}")
    
    generate_excel_report(results, best_config, target_ttft, target_tpot, args.output)


if __name__ == '__main__':
    main()